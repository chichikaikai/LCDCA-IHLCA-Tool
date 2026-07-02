"""
DB 上傳格式 Excel 匯出（強化版新增）
產出對齊網站資料庫 `products-import-template` 6 分頁格式
"""
import io
from datetime import datetime

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Sheet 1: products-import-template 的 28 個欄位（依使用者指定順序）──
PRODUCTS_HEADERS = [
    "PCCES編碼",            # 1
    "產品名稱",              # 2
    "產品英文名稱",          # 3
    "數量",                  # 4
    "宣告單位",              # 5
    "碳足跡數值",            # 6 ← 計算結果
    "製造地點",              # 7
    "系統邊界",              # 8
    "盤查起訖日",            # 9
    "製程描述",              # 10
    "排除項目",              # 11
    "LCA方法學",             # 12
    "活動數據來源",          # 13
    "排放係數來源",          # 14
    "數據品質等級可靠性",    # 15
    "數據品質等級完整性",    # 16
    "GWP方法",               # 17
    "建置單位",              # 18
    "查驗證說明",            # 19
    "建立資料時間",          # 20
    "更新資料時間",          # 21
    "版次",                  # 22 ← 輸出加「IH-」前綴
    "產品單價",              # 23 ← 移到後方
    "原物料投入",            # 24 ← 原名 T 拿掉
    "原物料單價",            # 25
    "原物料排放係數",        # 26 ← 原名 B 拿掉
    "數據品質",              # 27
    "碳足跡熱點",            # 28
]


# ────────────────────────────────────────────────────
def _box():
    s = Side(style="thin", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="2E4057")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _box()


def _style_input(cell):
    cell.font = Font(name="Arial", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _box()


# ────────────────────────────────────────────────────
def export_db_format(result, metadata, raw_rows, energy_rows,
                     quality_scores, calculator):
    """
    產出對齊 products-import-template 格式的 6 分頁 Excel

    metadata: dict 含 PCCES編碼、版次、計算方法、各項描述
    result   : core.calculator.LCAResult
    raw_rows / energy_rows : list of dicts (session_state.*)
    quality_scores : list of dicts (Re_act, Co_act, ..., Te_em)
    calculator : HybridLCACalculator（取 B_IO 和部門名稱用）

    回傳 BytesIO
    """
    wb = openpyxl.Workbook()

    # ============ Sheet 1: products-import-template ============
    ws1 = wb.active
    ws1.title = "products-import-template"
    for c, h in enumerate(PRODUCTS_HEADERS, start=1):
        cell = ws1.cell(1, c, h)
        _style_header(cell)
        ws1.column_dimensions[cell.column_letter].width = 16
    ws1.row_dimensions[1].height = 36

    # 數據品質等級：優先取後設資料的使用者輸入；沒有才回退到品質評分平均
    if "quality_reli" in metadata and metadata["quality_reli"] is not None:
        avg_re = int(metadata["quality_reli"])
    elif quality_scores:
        re_vals = [int(q.get("Re_act", 2) or 2) for q in quality_scores]
        avg_re = round(sum(re_vals) / len(re_vals), 1) if re_vals else 2
    else:
        avg_re = 2

    if "quality_comp" in metadata and metadata["quality_comp"] is not None:
        avg_co = int(metadata["quality_comp"])
    elif quality_scores:
        co_vals = [int(q.get("Co_act", 2) or 2) for q in quality_scores]
        avg_co = round(sum(co_vals) / len(co_vals), 1) if co_vals else 2
    else:
        avg_co = 2

    today = datetime.now().strftime("%Y-%m-%d")
    cf_value = round(float(result.total_emission), 6)

    # 版次自動加 IH- 前綴（Integrated Hybrid）
    ver_raw = (metadata.get("version", "V1.0") or "V1.0").strip()
    ver_out = ver_raw if ver_raw.startswith("IH-") else f"IH-{ver_raw}"

    row1_vals = [
        metadata.get("pcces", ""),                                        # 1  PCCES編碼
        result.product_name,                                              # 2  產品名稱
        metadata.get("product_en", ""),                                   # 3  產品英文名稱
        1,                                                                # 4  數量
        metadata.get("unit", ""),                                         # 5  宣告單位
        cf_value,                                                         # 6  碳足跡數值
        metadata.get("location", "台灣"),                                  # 7  製造地點
        metadata.get("boundary", "搖籃到大門"),                            # 8  系統邊界
        metadata.get("period", ""),                                       # 9  盤查起訖日
        metadata.get("process_desc", ""),                                 # 10 製程描述
        metadata.get("exclusions", "無"),                                  # 11 排除項目
        metadata.get("lca_method", ""),                                   # 12 LCA方法學
        metadata.get("activity_src", ""),                                 # 13 活動數據來源
        metadata.get("emission_src", ""),                                 # 14 排放係數來源
        avg_re,                                                           # 15 可靠性
        avg_co,                                                           # 16 完整性
        metadata.get("gwp", "IPCC AR6"),                                  # 17 GWP方法
        metadata.get("org", ""),                                          # 18 建置單位
        metadata.get("audit_note", ""),                                   # 19 查驗證說明
        today,                                                            # 20 建立資料時間
        today,                                                            # 21 更新資料時間
        ver_out,                                                          # 22 版次 IH-V1.0
        float(metadata.get("product_price", 0) or 0),                     # 23 產品單價
        "(見『原物料投入 T』分頁)",                                         # 24 原物料投入
        "(見『原物料單價Cu』分頁)",                                         # 25 原物料單價
        "(見『原物料排放係數B』分頁)",                                      # 26 原物料排放係數
        "(見『數據品質』分頁)",                                              # 27 數據品質
        "(見『碳足跡熱點』分頁)",                                            # 28 碳足跡熱點
    ]
    for c, v in enumerate(row1_vals, start=1):
        cell = ws1.cell(2, c, v)
        _style_input(cell)
    ws1.row_dimensions[2].height = 48

    # ============ Sheet 2: 原物料投入 T ============
    ws2 = wb.create_sheet("原物料投入 T")
    headers2 = ["投入產出項目", "單位", "數量", "來源"]
    for c, h in enumerate(headers2, start=1):
        _style_header(ws2.cell(1, c, h))
        ws2.column_dimensions[ws2.cell(1, c).column_letter].width = 20
    src_act = metadata.get("activity_src", "")
    r_idx = 2
    for r in raw_rows + energy_rows:
        if not r.get("name"):
            continue
        ws2.cell(r_idx, 1, r["name"])
        ws2.cell(r_idx, 2, r.get("unit", ""))
        ws2.cell(r_idx, 3, -abs(float(r.get("qty") or 0)))   # 投入為負
        ws2.cell(r_idx, 4, src_act)
        for c in range(1, 5):
            _style_input(ws2.cell(r_idx, c))
        r_idx += 1

    # ============ Sheet 3: 原物料單價 Cu ============
    ws3 = wb.create_sheet("原物料單價Cu")
    headers3 = ["投入產出項目", "單位", "單價", "來源"]
    for c, h in enumerate(headers3, start=1):
        _style_header(ws3.cell(1, c, h))
        ws3.column_dimensions[ws3.cell(1, c).column_letter].width = 20
    up_src = metadata.get("up_src", "工程會公共工程價格資料庫")
    r_idx = 2
    for r in raw_rows + energy_rows:
        if not r.get("name"):
            continue
        ws3.cell(r_idx, 1, r["name"])
        ws3.cell(r_idx, 2, r.get("unit", ""))
        ws3.cell(r_idx, 3, float(r.get("price") or 0))
        ws3.cell(r_idx, 4, up_src)
        for c in range(1, 5):
            _style_input(ws3.cell(r_idx, c))
        r_idx += 1

    # ============ Sheet 4: 原物料排放係數 B（含部門對應） ============
    ws4 = wb.create_sheet("原物料排放係數B")
    headers4 = ["投入產出項目", "宣告單位", "排放係數\n(kgCO2e/宣告單位)",
                "所屬部門編號", "所屬部門名稱"]
    for c, h in enumerate(headers4, start=1):
        _style_header(ws4.cell(1, c, h))
        ws4.column_dimensions[ws4.cell(1, c).column_letter].width = 22
    r_idx = 2
    for r in raw_rows + energy_rows:
        if not r.get("name"):
            continue
        sid = int(r.get("sector", 0)) + 1   # 1-based
        if calculator and 0 < sid <= len(calculator.B_IO):
            bio = float(calculator.B_IO[sid - 1])
            sec_name = (calculator.sector_names[sid - 1]
                        if sid - 1 < len(calculator.sector_names) else "")
        else:
            bio = 0.0
            sec_name = ""
        ws4.cell(r_idx, 1, r["name"])
        ws4.cell(r_idx, 2, r.get("unit", ""))
        ws4.cell(r_idx, 3, round(bio, 8))
        ws4.cell(r_idx, 4, f"{sid:03d}")
        ws4.cell(r_idx, 5, sec_name)
        for c in range(1, 6):
            _style_input(ws4.cell(r_idx, c))
        r_idx += 1

    # ============ Sheet 5: 數據品質（Pedigree 5×2） ============
    ws5 = wb.create_sheet("數據品質")
    # A 欄「投入產出項目」標題垂直跨第 1+2 列（先填 A1，再合併 A1:A2）
    _style_header(ws5.cell(1, 1, "投入產出項目"))
    ws5.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws5.column_dimensions['A'].width = 18
    # 第 1 列：分組標題
    ws5.cell(1, 2, "活動數據等級")
    ws5.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    ws5.cell(1, 7, "排放係數等級")
    ws5.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
    for c in [2, 7]:
        _style_header(ws5.cell(1, c))
    # 第 2 列：欄位名（B2 起）
    headers5_b = ["Re可靠性", "Co完整性", "Ti時間相關性", "Ge地理相關性", "Te技術相關性",
                  "Re可靠性", "Co完整性", "Ti時間相關性", "Ge地理相關性", "Te技術相關性"]
    for c_off, h in enumerate(headers5_b, start=2):
        _style_header(ws5.cell(2, c_off, h))
        ws5.column_dimensions[ws5.cell(2, c_off).column_letter].width = 14
    r_idx = 3
    for q in (quality_scores or []):
        ws5.cell(r_idx, 1, q.get("material", ""))
        ws5.cell(r_idx, 2, int(q.get("Re_act", 3) or 3))
        ws5.cell(r_idx, 3, int(q.get("Co_act", 3) or 3))
        ws5.cell(r_idx, 4, int(q.get("Ti_act", 3) or 3))
        ws5.cell(r_idx, 5, int(q.get("Ge_act", 3) or 3))
        ws5.cell(r_idx, 6, int(q.get("Te_act", 3) or 3))
        ws5.cell(r_idx, 7, int(q.get("Re_em", 3) or 3))
        ws5.cell(r_idx, 8, int(q.get("Co_em", 3) or 3))
        ws5.cell(r_idx, 9, int(q.get("Ti_em", 3) or 3))
        ws5.cell(r_idx, 10, int(q.get("Ge_em", 3) or 3))
        ws5.cell(r_idx, 11, int(q.get("Te_em", 3) or 3))
        for c in range(1, 12):
            _style_input(ws5.cell(r_idx, c))
        r_idx += 1

    # ============ Sheet 6: 碳足跡熱點（按類別分組）============
    ws6 = wb.create_sheet("碳足跡熱點")
    headers6 = ["類別", "投入產出項目", "碳足跡kgCO2e", "佔比(%)"]
    for c, h in enumerate(headers6, start=1):
        _style_header(ws6.cell(1, c, h))
        ws6.column_dimensions[ws6.cell(1, c).column_letter].width = 22

    # 拆出 raw / energy（依 n_raw；若無此屬性則假設全為原物料）
    n_raw = getattr(result, "n_raw", len(result.material_names))

    rows_hot = []
    # 1. 原物料投入（T 矩陣製程項）— 全數收錄
    for name, em in zip(result.material_names[:n_raw],
                        result.material_emissions[:n_raw]):
        rows_hot.append(("原物料投入", name + "製造", float(em)))
    # 2. 能資源投入（T 矩陣製程項）— 全數收錄
    for name, em in zip(result.material_names[n_raw:],
                        result.material_emissions[n_raw:]):
        rows_hot.append(("能資源投入", name + "製造", float(em)))
    # 3. 產品（T 矩陣製程項）
    rows_hot.append(("產品", result.product_name + "製造",
                     float(result.product_emission)))
    # 4. IO 部門 — 非零項按絕對值降冪
    io_rows = []
    for name, em in zip(result.sector_names, result.sector_emissions):
        if abs(float(em)) > 1e-12:
            io_rows.append(("IO 部門", name, float(em)))
    io_rows.sort(key=lambda x: abs(x[2]), reverse=True)
    rows_hot.extend(io_rows)

    total_abs = sum(abs(e) for _, _, e in rows_hot) or 1.0

    r_idx = 2
    for cat, name, em in rows_hot:
        ws6.cell(r_idx, 1, cat)
        ws6.cell(r_idx, 2, name)
        ws6.cell(r_idx, 3, round(em, 6))
        ws6.cell(r_idx, 4, round(em / total_abs * 100, 2))
        for c in range(1, 5):
            _style_input(ws6.cell(r_idx, c))
        r_idx += 1

    # ── 輸出 ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_coefficient_id(pcces: str, method: str, version: str) -> str:
    """
    產出完整係數編號：PCCES-IH-版次
    方法欄固定為 IH（Integrated Hybrid），不受 method 參數影響
    例：M0337718003-IH-V1.0
    """
    pcces = (pcces or "").strip() or "UNKNOWN"
    version = (version or "V1.0").strip()
    return f"{pcces}-IH-{version}"
