import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import json
from pathlib import Path
import io


def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def make_header_style():
    return {
        "font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="2E4057"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": make_border(),
    }

def make_note_style():
    return {
        "font": Font(name="Arial", italic=True, color="595959", size=10),
        "fill": PatternFill("solid", fgColor="F2F2F2"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

def make_input_style(color="FFFFFF"):
    return {
        "font": Font(name="Arial", size=11),
        "fill": PatternFill("solid", fgColor=color),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": make_border("thin"),
    }

def make_type_style(fgColor):
    return {
        "font": Font(name="Arial", bold=True, size=11, color="2E4057"),
        "fill": PatternFill("solid", fgColor=fgColor),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": make_border(),
    }

def apply_style(cell, style_dict):
    for k, v in style_dict.items():
        setattr(cell, k, v)


def _load_sector_names():
    """讀取內建 163 部門名稱"""
    data_dir = Path(__file__).parent.parent / "data"
    try:
        with open(data_dir / "tw110_sectors.json", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return [f"{i+1:03d}" for i in range(163)]


# ── Bᴀ 係數向量範本（強化版：直欄填表，A=編號 B=名稱 C=數值）─
def create_BA_template() -> bytes:
    """產出單一分頁 Bᴀ 的範本 Excel；使用者只填 C 欄 163 個數值"""
    sector_names = _load_sector_names()
    n = len(sector_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B_A"

    # 表頭（第 1 列）
    headers = ["部門編號", "部門名稱", "排放係數 (kgCO₂e/元)"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(1, j, h)
        apply_style(cell, make_header_style())
    ws.row_dimensions[1].height = 36
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22

    # 凍結首列
    ws.freeze_panes = "A2"

    # 第 2~164 列：163 個部門資料
    for i, name in enumerate(sector_names, start=2):
        # 拆出編號與名稱（內建格式可能是「001稻作」也可能就是「稻作」）
        if len(name) >= 3 and name[:3].isdigit():
            num, sname = name[:3], name[3:]
        else:
            num, sname = f"{i-1:03d}", name

        c_num = ws.cell(i, 1, num)
        c_num.font = Font(name="Arial", size=11, bold=True, color="595959")
        c_num.alignment = Alignment(horizontal="center", vertical="center")
        c_num.border = make_border("thin")
        c_num.fill = PatternFill("solid", fgColor="F2F2F2")

        c_name = ws.cell(i, 2, sname)
        c_name.font = Font(name="Microsoft JhengHei", size=11)
        c_name.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_name.border = make_border("thin")
        c_name.fill = PatternFill("solid", fgColor="F2F2F2")

        # C 欄：使用者填值（白底、可輸入樣式）
        c_val = ws.cell(i, 3, "")
        apply_style(c_val, make_input_style())
        c_val.number_format = "0.000000"
        ws.row_dimensions[i].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── IO 資料範本（舊版，保留供向後相容）─────────────────────
def create_io_template() -> bytes:
    wb = openpyxl.Workbook()

    # IO_A
    ws_a = wb.active
    ws_a.title = "IO_A"
    ws_a.merge_cells("A1:D1")
    n = ws_a["A1"]
    n.value = "【IO_A】請從 A2 開始貼入 A 矩陣（sectorCount × sectorCount），不需表頭"
    apply_style(n, make_note_style())
    ws_a.row_dimensions[1].height = 30
    ws_a.column_dimensions["A"].width = 60

    # B_IO
    ws_b = wb.create_sheet("B_IO")
    ws_b.merge_cells("A1:D1")
    n = ws_b["A1"]
    n.value = "【B_IO】請從 A2 開始橫向貼入碳強度向量（1 列 × sectorCount 欄），不需表頭"
    apply_style(n, make_note_style())
    ws_b.row_dimensions[1].height = 30
    ws_b.column_dimensions["A"].width = 60

    # Sector
    ws_sec = wb.create_sheet("Sector")
    ws_sec.merge_cells("A1:B1")
    n = ws_sec["A1"]
    n.value = "【Sector】請從 A2 開始直向貼入部門名稱，順序須與 IO_A 及 B_IO 一致"
    apply_style(n, make_note_style())
    ws_sec.row_dimensions[1].height = 30
    ws_sec.column_dimensions["A"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UserInput 範本（含部門下拉、類型含產品）────────────────
def create_userinput_template() -> bytes:
    sector_names = _load_sector_names()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UserInput"

    # 說明列
    ws.merge_cells("A1:F1")
    note = ws["A1"]
    note.value = "【投入產出填報表】類型：產品 / 原物料 / 能資源；對應部門請從下拉選單選取"
    apply_style(note, make_note_style())
    ws.row_dimensions[1].height = 30

    # 表頭（含部門欄）
    headers    = ["類型", "名稱", "對應 IO 部門", "單位", "數量", "單價\n($NTD/單位)"]
    col_widths = [12,     16,     30,              8,     10,     18]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        apply_style(cell, make_header_style())
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 36

    # 類型下拉（A欄）
    dv_type = DataValidation(
        type="list",
        formula1='"產品,原物料,能資源"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.error = "請選擇「產品」、「原物料」或「能資源」"
    dv_type.errorTitle = "輸入錯誤"
    ws.add_data_validation(dv_type)
    dv_type.add("A3:A500")

    # 部門下拉（C欄）— 用隱藏工作表存放部門清單
    ws_sec = wb.create_sheet("_Sectors", -1)
    ws_sec.sheet_state = "hidden"
    for i, name in enumerate(sector_names, 1):
        ws_sec.cell(row=i, column=1, value=name)

    sec_ref = f"_Sectors!$A$1:$A${len(sector_names)}"
    dv_sec = DataValidation(
        type="list",
        formula1=sec_ref,
        allow_blank=True,
        showDropDown=False,
    )
    dv_sec.error = "請從清單選取對應部門"
    dv_sec.errorTitle = "輸入錯誤"
    ws.add_data_validation(dv_sec)
    dv_sec.add("C3:C500")

    # 填色設定
    prod_fill  = "D6EAF8"   # 淡藍，產品
    raw_fill   = "EAF3DE"   # 淡綠，原物料
    eng_fill   = "FFF3CD"   # 淡黃，能資源
    blank_fill = "FAFAFA"

    examples = [
        ("產品",   "預拌混凝土", "065水泥製品",       "M3",  1,    3027.0),
        ("原物料", "砂",         "013砂、石及其他礦產品", "kg",  989,  0.37),
        ("原物料", "石",         "013砂、石及其他礦產品", "kg",  850,  0.41),
        ("原物料", "水泥",       "064水泥",            "kg",  148,  2.86),
        ("原物料", "爐石粉",     "122其他陸上運輸",     "kg",  98,   1.30),
        ("原物料", "飛灰",       "122其他陸上運輸",     "kg",  43,   0.46),
        ("原物料", "矽灰",       "122其他陸上運輸",     "kg",  0,    0.00),
        ("原物料", "減水劑",     "058未分類其他化學製品","kg",  2.9,  16.0),
        ("原物料", "水",         "109自來水",           "kg",  185,  0.01),
        ("能資源", "電力",       "107電力及蒸汽",       "kWh", 2,    2.63),
    ]

    fill_map = {"產品": prod_fill, "原物料": raw_fill, "能資源": eng_fill}

    for r, row_data in enumerate(examples, 3):
        mat_type = row_data[0]
        fc = fill_map.get(mat_type, blank_fill)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                apply_style(cell, make_type_style(fc))
            else:
                apply_style(cell, make_input_style(fc))

    # 空白填報列
    for r in range(len(examples) + 3, len(examples) + 23):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c, value="")
            apply_style(cell, make_input_style(blank_fill))

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# 相容舊呼叫
def create_template() -> bytes:
    return create_io_template()
